"""Bulk-adding products from an admin-supplied Excel sheet.

Validated in full before anything is written: a sheet with one bad row would
otherwise leave a half-imported catalog with no clean way to tell what
succeeded. Sellers are the one thing this creates on the fly, by name --
that is the point of the feature. Units are not: the shop's unit list is
deliberately exact, so an unrecognised one fails the whole sheet instead of
minting another.
"""

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .models import Product, Seller, Unit

TEMPLATE_HEADERS = ["Name", "Order name", "Seller", "Unit", "Price"]


@dataclass
class ImportRow:
    row_number: int
    name: str
    order_name: str
    seller_name: str
    unit: Unit
    price: Decimal | None


@dataclass
class RowError:
    row_number: int
    message: str

    def __str__(self) -> str:
        return f"Row {self.row_number}: {self.message}"


@dataclass
class ImportResult:
    created_count: int
    sellers_created_count: int
    skipped_count: int


class InvalidWorkbook(Exception):
    """The uploaded file could not be read as an .xlsx workbook at all."""


def build_template_workbook() -> Workbook:
    """The downloadable starting point: headers, one example row, and a
    second sheet listing the units the first sheet's Unit column has to
    match -- exact-match only, so the admin needs to see the real list
    rather than guess at spelling.
    """
    wb = Workbook()

    products = wb.active
    products.title = "Products"
    products.append(TEMPLATE_HEADERS)
    for cell in products[1]:
        cell.font = Font(bold=True)
    products.append(["Whole milk", "", "Green Valley Dairy", "τμχ", "1.15"])
    for column, width in zip("ABCDE", (28, 20, 24, 14, 10)):
        products.column_dimensions[column].width = width

    units = wb.create_sheet("Valid units")
    units.append(["Name", "Plural"])
    for cell in units[1]:
        cell.font = Font(bold=True)
    for row, unit in enumerate(Unit.objects.all(), start=2):
        units.cell(row=row, column=1, value=unit.name)
        units.cell(row=row, column=2, value=unit.plural)
    for column, width in zip("AB", (20, 20)):
        units.column_dimensions[column].width = width

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _clean_price(raw, row_number: int, errors: list[RowError]) -> Decimal | None:
    if raw is None or str(raw).strip() == "":
        return None
    try:
        price = Decimal(str(raw).strip())
    except InvalidOperation:
        errors.append(RowError(row_number, f"Price “{raw}” is not a valid number."))
        return None
    if price <= 0:
        errors.append(RowError(row_number, "Price must be more than zero, or left empty."))
        return None
    return price


def parse_workbook(file) -> tuple[list[ImportRow], list[RowError]]:
    """Read and validate every row before anything touches the database.

    Blank rows (a trailing empty row is the usual cause) are skipped rather
    than reported -- there is nothing wrong to fix.
    """
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception as exc:
        raise InvalidWorkbook(str(exc)) from exc

    sheet = wb.worksheets[0]
    units_by_name = {unit.name.strip().lower(): unit for unit in Unit.objects.all()}

    rows: list[ImportRow] = []
    errors: list[RowError] = []

    for row_number, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        name, order_name, seller_name, unit_name, price_raw = (list(cells) + [None] * 5)[:5]
        if all(cell is None or str(cell).strip() == "" for cell in (name, seller_name, unit_name, price_raw)):
            continue

        row_errors: list[RowError] = []

        name = (name or "").strip() if name is not None else ""
        if not name:
            row_errors.append(RowError(row_number, "Name is required."))

        seller_name = (str(seller_name).strip() if seller_name is not None else "")
        if not seller_name:
            row_errors.append(RowError(row_number, "Seller is required."))

        unit_name = (str(unit_name).strip() if unit_name is not None else "")
        unit = units_by_name.get(unit_name.lower())
        if not unit_name:
            row_errors.append(RowError(row_number, "Unit is required."))
        elif unit is None:
            row_errors.append(
                RowError(row_number, f"“{unit_name}” is not one of the shop's units.")
            )

        price = _clean_price(price_raw, row_number, row_errors)

        if row_errors:
            errors.extend(row_errors)
            continue

        rows.append(ImportRow(
            row_number=row_number,
            name=name,
            order_name=(str(order_name).strip() if order_name is not None else ""),
            seller_name=seller_name,
            unit=unit,
            price=price,
        ))

    return rows, errors


def import_products(rows: list[ImportRow], created_by) -> ImportResult:
    """Commit a sheet that has already passed `parse_workbook` clean.

    Sellers are matched case-insensitively, same as everywhere else a seller
    name is typed by hand, and created on the fly when new. A row matching a
    product that already exists for that seller -- including one just
    created earlier in this same sheet -- is skipped rather than touched,
    so the same sheet can be re-uploaded safely after fixing only some rows.
    """
    created_count = 0
    sellers_created_count = 0
    skipped_count = 0

    with transaction.atomic():
        for row in rows:
            seller = Seller.objects.filter(name__iexact=row.seller_name).first()
            if seller is None:
                seller = Seller.objects.create(name=row.seller_name)
                sellers_created_count += 1

            if Product.objects.filter(seller=seller, name__iexact=row.name).exists():
                skipped_count += 1
                continue

            Product.objects.create(
                name=row.name,
                order_name=row.order_name,
                seller=seller,
                unit=row.unit,
                unit_price=row.price,
                created_by=created_by,
            )
            created_count += 1

    return ImportResult(
        created_count=created_count,
        sellers_created_count=sellers_created_count,
        skipped_count=skipped_count,
    )
