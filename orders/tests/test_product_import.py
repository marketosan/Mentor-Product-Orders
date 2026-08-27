"""Bulk-importing products from an admin-supplied Excel sheet.

A sheet with any bad row is rejected whole -- nothing is written until every
row is clean. Sellers are matched by name and created on the fly; units are
matched only, never created, since the shop's unit list is deliberately exact.
"""

import json
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from orders.models import Product, Seller, Unit
from orders.product_import import ImportRow, RowError, import_products, parse_workbook

User = get_user_model()

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def sheet_file(rows, filename="import.xlsx"):
    """An in-memory .xlsx, header row plus whatever data rows are given."""
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Order name", "Seller", "Unit", "Price"])
    for row in rows:
        ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    return SimpleUploadedFile(filename, buffer.getvalue(), content_type=XLSX_CONTENT_TYPE)


class ProductImportTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.admin = User.objects.create_user(
            username="boss", email="boss@example.com", password="pw", role=User.Role.ADMIN
        )
        cls.maria = User.objects.create_user(
            username="maria", email="maria@example.com", password="pw"
        )
        cls.dairy = Seller.objects.create(name="Green Valley Dairy")
        # Seeded by the 0004_unit migration, which runs on the test database
        # too -- not created here, since the unit list is meant to be fixed.
        cls.litre = Unit.objects.get(name="τμχ")

    def setUp(self):
        self.client.force_login(self.admin)

    def toast(self, response):
        return json.loads(response.headers["HX-Trigger"])["toast"]["message"]


class ParseWorkbookTests(ProductImportTestCase):
    def test_a_clean_row_parses(self):
        rows, errors = parse_workbook(
            sheet_file([["Whole milk", "MILK-1L", "Green Valley Dairy", "τμχ", "1.15"]])
        )

        self.assertEqual(errors, [])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].name, "Whole milk")
        self.assertEqual(rows[0].order_name, "MILK-1L")
        self.assertEqual(rows[0].seller_name, "Green Valley Dairy")
        self.assertEqual(rows[0].unit, self.litre)
        self.assertEqual(rows[0].price, Decimal("1.15"))

    def test_a_blank_price_is_none_not_an_error(self):
        rows, errors = parse_workbook(
            sheet_file([["Whole milk", "", "Green Valley Dairy", "τμχ", None]])
        )

        self.assertEqual(errors, [])
        self.assertIsNone(rows[0].price)

    def test_a_missing_name_is_an_error(self):
        rows, errors = parse_workbook(
            sheet_file([["", "", "Green Valley Dairy", "τμχ", "1.15"]])
        )

        self.assertEqual(rows, [])
        self.assertIn("Name is required", str(errors[0]))
        self.assertTrue(str(errors[0]).startswith("Row 2:"))

    def test_a_missing_seller_is_an_error(self):
        rows, errors = parse_workbook(
            sheet_file([["Whole milk", "", "", "τμχ", "1.15"]])
        )

        self.assertIn("Seller is required", str(errors[0]))

    def test_an_unrecognised_unit_is_an_error_not_a_new_unit(self):
        rows, errors = parse_workbook(
            sheet_file([["Whole milk", "", "Green Valley Dairy", "λίτρο", "1.15"]])
        )

        self.assertEqual(rows, [])
        self.assertIn("not one of the shop's units", str(errors[0]))
        self.assertFalse(Unit.objects.filter(name="λίτρο").exists())

    def test_unit_matching_ignores_case(self):
        rows, errors = parse_workbook(
            sheet_file([["Whole milk", "", "Green Valley Dairy", "ΤΜΧ", "1.15"]])
        )

        self.assertEqual(errors, [])
        self.assertEqual(rows[0].unit, self.litre)

    def test_a_non_numeric_price_is_an_error(self):
        rows, errors = parse_workbook(
            sheet_file([["Whole milk", "", "Green Valley Dairy", "τμχ", "free"]])
        )

        self.assertIn("not a valid number", str(errors[0]))

    def test_a_zero_price_is_an_error(self):
        rows, errors = parse_workbook(
            sheet_file([["Whole milk", "", "Green Valley Dairy", "τμχ", "0"]])
        )

        self.assertIn("more than zero", str(errors[0]))

    def test_a_wholly_blank_row_is_skipped_silently(self):
        rows, errors = parse_workbook(sheet_file([
            ["Whole milk", "", "Green Valley Dairy", "τμχ", "1.15"],
            [None, None, None, None, None],
        ]))

        self.assertEqual(errors, [])
        self.assertEqual(len(rows), 1)

    def test_row_numbers_match_the_spreadsheet(self):
        """Row 1 is the header, so the first data row is row 2."""
        rows, errors = parse_workbook(sheet_file([
            ["Whole milk", "", "Green Valley Dairy", "τμχ", "1.15"],
            ["", "", "Green Valley Dairy", "τμχ", "1.15"],
        ]))

        self.assertTrue(str(errors[0]).startswith("Row 3:"))


class ImportProductsTests(ProductImportTestCase):
    def row(self, **overrides):
        defaults = {
            "row_number": 2, "name": "Whole milk", "order_name": "",
            "seller_name": "Green Valley Dairy", "unit": self.litre, "price": Decimal("1.15"),
        }
        defaults.update(overrides)
        return ImportRow(**defaults)

    def test_it_creates_a_product_under_an_existing_seller(self):
        result = import_products([self.row()], created_by=self.admin)

        product = Product.objects.get(name="Whole milk")
        self.assertEqual(product.seller, self.dairy)
        self.assertEqual(product.created_by, self.admin)
        self.assertEqual(result.created_count, 1)
        self.assertEqual(result.sellers_created_count, 0)

    def test_an_unknown_seller_is_created_by_name(self):
        result = import_products([self.row(seller_name="Corner Shop")], created_by=self.admin)

        self.assertTrue(Seller.objects.filter(name="Corner Shop").exists())
        self.assertEqual(result.sellers_created_count, 1)

    def test_seller_matching_ignores_case(self):
        import_products([self.row(seller_name="green valley dairy")], created_by=self.admin)

        self.assertEqual(Seller.objects.filter(name__iexact="green valley dairy").count(), 1)
        self.assertEqual(Product.objects.get(name="Whole milk").seller, self.dairy)

    def test_a_product_already_on_the_catalog_is_skipped(self):
        Product.objects.create(
            name="Whole milk", seller=self.dairy, unit=self.litre, unit_price=Decimal("1.00")
        )

        result = import_products([self.row()], created_by=self.admin)

        self.assertEqual(result.created_count, 0)
        self.assertEqual(result.skipped_count, 1)
        # The pre-existing price is untouched -- this is a skip, not an update.
        self.assertEqual(Product.objects.get(name="Whole milk").unit_price, Decimal("1.00"))

    def test_a_duplicate_within_the_same_sheet_is_skipped_once(self):
        result = import_products([self.row(), self.row(row_number=3)], created_by=self.admin)

        self.assertEqual(Product.objects.filter(name="Whole milk").count(), 1)
        self.assertEqual(result.created_count, 1)
        self.assertEqual(result.skipped_count, 1)

    def test_an_optional_price_is_left_unknown(self):
        import_products([self.row(price=None)], created_by=self.admin)

        self.assertIsNone(Product.objects.get(name="Whole milk").unit_price)


class ImportViewTests(ProductImportTestCase):
    def test_the_button_opens_the_dialog(self):
        response = self.client.get(reverse("import_products"))

        self.assertTemplateUsed(response, "orders/_product_import_form.html")
        self.assertContains(response, "Import products")

    def test_a_clean_sheet_imports_and_confirms(self):
        response = self.client.post(reverse("import_products"), {
            "file": sheet_file([["Whole milk", "", "Green Valley Dairy", "τμχ", "1.15"]]),
        })

        self.assertTrue(Product.objects.filter(name="Whole milk").exists())
        self.assertEqual(response.headers["HX-Retarget"], "#product-list")
        fired = json.loads(response.headers["HX-Trigger"])
        self.assertTrue(fired["closeModal"])
        self.assertIn("1 added", fired["toast"]["message"])

    def test_the_toast_mentions_sellers_created_and_rows_skipped(self):
        Product.objects.create(
            name="Napkins", seller=self.dairy, unit=self.litre, unit_price=Decimal("8.75")
        )

        response = self.client.post(reverse("import_products"), {
            "file": sheet_file([
                ["Whole milk", "", "Corner Shop", "τμχ", "1.15"],
                ["Napkins", "", "Green Valley Dairy", "τμχ", "8.75"],
            ]),
        })

        message = self.toast(response)
        self.assertIn("1 added", message)
        self.assertIn("1 seller created", message)
        self.assertIn("1 already on the catalog, skipped", message)

    def test_a_bad_row_imports_nothing_and_lists_the_problem(self):
        response = self.client.post(reverse("import_products"), {
            "file": sheet_file([
                ["Whole milk", "", "Green Valley Dairy", "τμχ", "1.15"],
                ["", "", "Green Valley Dairy", "τμχ", "1.15"],
            ]),
        })

        self.assertEqual(response.status_code, 422)
        self.assertContains(response, "Row 3", status_code=422)
        self.assertContains(response, "Name is required", status_code=422)
        self.assertFalse(Product.objects.filter(name="Whole milk").exists())

    def test_an_unreadable_file_is_refused_cleanly(self):
        bogus = SimpleUploadedFile("import.xlsx", b"not a spreadsheet", content_type=XLSX_CONTENT_TYPE)

        response = self.client.post(reverse("import_products"), {"file": bogus})

        self.assertEqual(response.status_code, 422)
        self.assertContains(response, "valid .xlsx file", status_code=422)

    def test_a_missing_file_is_refused_cleanly(self):
        response = self.client.post(reverse("import_products"), {})

        self.assertEqual(response.status_code, 422)
        self.assertContains(response, "Choose a file", status_code=422)

    def test_a_sheet_with_no_rows_is_refused_cleanly(self):
        response = self.client.post(reverse("import_products"), {"file": sheet_file([])})

        self.assertEqual(response.status_code, 422)
        self.assertContains(response, "no product rows", status_code=422)

    def test_an_employee_cannot_import(self):
        self.client.force_login(self.maria)

        response = self.client.post(reverse("import_products"), {
            "file": sheet_file([["Whole milk", "", "Green Valley Dairy", "τμχ", "1.15"]]),
        })

        self.assertEqual(response.status_code, 403)
        self.assertFalse(Product.objects.filter(name="Whole milk").exists())


class TemplateDownloadTests(ProductImportTestCase):
    def test_it_downloads_an_xlsx_file(self):
        response = self.client.get(reverse("product_import_template"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["Content-Type"],
            XLSX_CONTENT_TYPE,
        )
        self.assertIn("products-template.xlsx", response.headers["Content-Disposition"])

    def test_it_has_the_expected_headers_and_a_units_reference_sheet(self):
        response = self.client.get(reverse("product_import_template"))

        wb = load_workbook(BytesIO(response.content))
        products = wb["Products"]
        self.assertEqual(
            [cell.value for cell in products[1]],
            ["Name", "Order name", "Seller", "Unit", "Price"],
        )

        units_sheet = wb["Valid units"]
        names = [row[0] for row in units_sheet.iter_rows(min_row=2, values_only=True)]
        self.assertIn("τμχ", names)

    def test_an_employee_cannot_download_it(self):
        self.client.force_login(self.maria)

        self.assertEqual(self.client.get(reverse("product_import_template")).status_code, 403)
